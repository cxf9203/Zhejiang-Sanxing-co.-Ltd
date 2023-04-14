# update : 2022.06.30
#################################导入必备的库（工具）#############################
# version:1.0.2022.0630
import gxipy as gx
from scipy.optimize import leastsq
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog
import sanxing08
import zichuangkou03
import labelinfo01
import procCallback
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5 import QtWidgets, QtCore, QtGui
import cv2
import time
import qrcode
import win32print#windows打印机
import win32ui
import win32api
from PIL import Image, ImageWin, ImageDraw, ImageFont
import numpy as np
import threading
import os
import openpyxl
from openpyxl.drawing.image import Image as IMG
import math
import datetime

################################初始化参数设置
shuzu = list()
camera1k, camera1b = 0, 0
camera2k, camera2b = 0, 0
camera3k, camera3b = 0, 0
camera4k, camera4b = 0, 0
Year = {"22": "T", "23": "W", "24": "X"}
Month = {"01": "1", "02": "2", "03": "3", "04": "4", "05": "5",
         "06": "6", "07": "7", "08": "8", "09": "9",
         "10": "A", "11": "B", "12": "C"}
Day = {"01": "1", "02": "2", "03": "3", "04": "4", "05": "5", "06": "6", "07": "7", "08": "8", "09": "9",
       "10": "A", "11": "B", "12": "C", "13": "D",
       "14": "E", "15": "F", "16": "G", "17": "H", "18": "I", "19": "J", "20": "K", "21": "L", "22": "M", "23": "N",
       "24": "O", "25": "P",
       "26": "Q", "27": "R", "28": "S", "29": "T", "30": "U", "31": "V"}

################################阈值滑条设置###################
alpha_slider_max = 255
#img= cv2.imread("test.jpg",0)
def on_trackbar(val):
    global  img, yuzhi
    yuzhi = val
    # dst = cv.cvtColor(src1,cv.COLOR_BGR2GRAY)
    ret, binary = cv2.threshold(img,yuzhi, 255, cv2.THRESH_BINARY)
    cv2.imshow("CANVAS", binary)
#######################手动创建ROI###########################
# 创建图像与窗口并将窗口与回调函数绑定
def create_roi_automatically(numpy_image, flag=0):
    global img,yuzhi
    img = numpy_image
    if flag==2:
        img = cv2.resize(img, (int(img.shape[1]), int(img.shape[0])))
    elif flag==1:
        img = cv2.resize(img, (int(img.shape[1]/3), int(img.shape[0]/3)))
    else:
        img = cv2.resize(img, (int(img.shape[1] / 2), int(img.shape[0] / 2)))
    cv2.namedWindow('CANVAS')
    cv2.setMouseCallback('CANVAS', on_mouse)
    trackbar_name = 'Alpha x %d' % alpha_slider_max
    cv2.createTrackbar(trackbar_name, "CANVAS", 50, alpha_slider_max, on_trackbar)
    # Show some stuff
    on_trackbar(0)
    # Wait until user press some key
    cv2.waitKey(0)
    print(yuzhi)
    cv2.destroyAllWindows()



def on_mouse(event, x, y, flags, param):
    global img, point1, point2, g_rect
    img2 = img.copy()
    if event == cv2.EVENT_LBUTTONDOWN:  # 左键点击,则在原图打点
        print("1-EVENT_LBUTTONDOWN")
        point1 = (x, y)
        cv2.circle(img2, point1, 10, (0, 255, 0), 5)
        cv2.imshow('CANVAS', img2)
        print(point1)

    elif event == cv2.EVENT_MOUSEMOVE and (flags & cv2.EVENT_FLAG_LBUTTON):  # 按住左键拖曳，画框
        print("2-EVENT_FLAG_LBUTTON")
        cv2.rectangle(img2, point1, (x, y), (255, 0, 0), thickness=2)
        cv2.imshow('CANVAS', img2)
        print(x, y)

    elif event == cv2.EVENT_LBUTTONUP:  # 左键释放，显示
        print("3-EVENT_LBUTTONUP")
        point2 = (x, y)
        cv2.rectangle(img2, point1, point2, (0, 0, 255), thickness=2)
        cv2.imshow('CANVAS', img2)
        """
        if point1 != point2:
            min_x = min(point1[0], point2[0])
            min_y = min(point1[1], point2[1])
            width = abs(point1[0] - point2[0])
            height = abs(point1[1] - point2[1])
            g_rect = [min_x, min_y, width, height]
            cut_img = img[min_y:min_y + height, min_x:min_x + width]
            cv2.imshow('ROI', cut_img)
            """

#############################定义一个数据保存函数，能被MES系统查看############
def data_record(information):
    info = information
    # path = "record"
    # os.chdir(path)  # 修改工作路径
    try:
        workbook = openpyxl.load_workbook('data_record.xlsx')  # 返回一个workbook数据类型的值.
    except:
        wb = openpyxl.Workbook()
        wb.save('data_record.xlsx')
        workbook = openpyxl.load_workbook('data_record.xlsx')
        sheet = workbook.active  # 获取活动表
        sheet['A1'] = "日期"
        workbook['A2'] = datetime.datetime.now()
        sheet['B1'] = "信息"
        sheet['C1'] = "操作员名字"
        workbook.save('data_record.xlsx')
    sheet = workbook.active  # 获取活动表
    rows = sheet.max_row  # 获取最大行数
    print(rows)
    # data = sheet.cell(row=3, column=1)
    # print(data.value)
    sheet['B' + str(rows + 1)] = info  # 在该行新增标签信息
    sheet['A' + str(rows + 1)] = datetime.datetime.now()  # 在改行新增操作日期
    workbook.save('data_record.xlsx')  # 保存文档


####################################################子窗口类######################################
class labelwindow(QMainWindow, labelinfo01.Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        labelinfo01.Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.comboBox.currentIndexChanged.connect(self.liheqixuanxing)
        self.lineEdit_45.editingFinished.connect(self.zubie_edit)
        self.lineEdit_5.editingFinished.connect(self.ProductLine_edit)
        self.comboBox_2.currentIndexChanged.connect(self.fontstyle)
        self.pushButton.clicked.connect(self.serial_number_clear)
    def serial_number_clear(self):
        print("clear serial number")
        md.serial_number = 0
        md.sheet1['B1'] = md.serial_number#save srial number

    def fontstyle(self, i):
        print("Items in the list are :")
        for count in range(self.comboBox_2.count()):
            print(self.comboBox_2.itemText(count))  # Displays text belonging to specific index
        print("Current index", i, "selection changed ", self.comboBox_2.currentText())
        if i == 0:
            print("choose 宋体")
            md.fonts = 'C:/Windows/Fonts/STFANGSO.TTF'
        elif i == 1:
            print("choose 黑体")
            md.fonts = 'C:/Windows/Fonts/STFANGSO.TTF'
        elif i == 2:
            print("choose 粗体")
            md.fonts = 'C:/Windows/Fonts/ARLRDBD.TTF'

    def zubie_edit(self):
        md.zubie = self.lineEdit_45.text()
        md.sheet1['B11'] = md.zubie
        print("parameter get")
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")
    def ProductLine_edit(self):
        md.scxian = self.lineEdit_5.text()
        md.sheet1['B56'] = md.scxian
        print("parameter get")
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")
    def liheqixuanxing(self, i):
        print("Items in the list are :")
        for count in range(self.comboBox.count()):
            print(self.comboBox.itemText(count))  # Displays text belonging to specific index
        print("Current index", i, "selection changed ", self.comboBox.currentText())
        if i == 0:
            print("choose N-15.0离合器")
            self.liheqi = int(0)
            #todo 参数自动生成
            self.lineEdit_23.setText("A158") #项目代码
            self.lineEdit_28.setText("6.00")#上限规格
            self.lineEdit_21.setText("3.00")#下限规格
            self.lineEdit_22.setText("A156")  # 项目代码
            self.lineEdit_20.setText("58.00")  # 上限规格
            self.lineEdit_19.setText("57.00")  # 下限规格
            self.lineEdit_44.setText("A209")  # 项目代码
            self.lineEdit_42.setText("43.90")  # 上限规格
            self.lineEdit_31.setText("42.90")  # 下限规格
            self.lineEdit_2.setText("DC9716984A")  # 
            md.sheet1['B51'] = self.liheqi  # 离合器选型保存等待下次加载
            md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        elif i == 1:
            print("choose N-5.25离合器")
            self.liheqi = int(1)
            # todo 参数自动生成
            self.lineEdit_23.setText("A162")  # 项目代码
            self.lineEdit_28.setText("50.00")  # 上限规格
            self.lineEdit_21.setText("47.00")  # 下限规格
            self.lineEdit_22.setText("A209")  # 项目代码
            self.lineEdit_20.setText("46.70")  # 上限规格
            self.lineEdit_19.setText("45.70")  # 下限规格
            self.lineEdit_44.setText("A156")  # 项目代码
            self.lineEdit_42.setText("53.00")  # 上限规格
            self.lineEdit_31.setText("52.00")  # 下限规格
            self.lineEdit_2.setText("DC9111644E")  # 
            md.sheet1['B51'] = self.liheqi  # 离合器选型保存等待下次加载
            md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        elif i == 2:
            print("choose NBP-11.0离合器")
            self.liheqi = int(2)
            # todo 参数自动生成
            self.lineEdit_23.setText("A162")  # 项目代码
            self.lineEdit_28.setText("50.00")  # 上限规格
            self.lineEdit_21.setText("47.00")  # 下限规格
            self.lineEdit_22.setText("A158")  # 项目代码
            self.lineEdit_20.setText("6.00")  # 上限规格
            self.lineEdit_19.setText("3.00")  # 下限规格
            self.lineEdit_44.setText("A156")  # 项目代码
            self.lineEdit_42.setText("53.00")  # 上限规格
            self.lineEdit_31.setText("52.00")  # 下限规格
            self.lineEdit_2.setText("DC9721134A")  # 
            md.sheet1['B51'] = self.liheqi  # 离合器选型保存等待下次加载
            md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
#######################################################子窗口类#######################################
class biaodingwindow(QMainWindow, zichuangkou03.Ui_MainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        zichuangkou03.Ui_MainWindow.__init__(self)
        self.setupUi(self)
        self.camera_califlag = 0
        # self.painter = QPainter(self)
        self.camera1k = 0
        self.camera1b = 0
        self.camera2k = 0
        self.camera2b = 0
        self.camera3k = 0
        self.camera3b = 0
        self.ccdyuzhi = [0,0]
        self.ccd2yuzhi = [0,0]
        self.ccd3yuzhi = [0,0]
        self.cam3caiji = False
        self.pushButton.clicked.connect(self.camera_cali)
        self.pushButton_2.clicked.connect(self.zichuangkouclose)
        self.pushButton_3.clicked.connect(self.camera_cali2)
        self.pushButton_4.clicked.connect(self.camera_cali3)
        self.pushButton_6.clicked.connect(self.canshuqueren)
        self.pushButton_7.clicked.connect(self.save_parameter)
        self.pushButton_8.clicked.connect(self.load_parameter)
        self.horizontalSlider.sliderReleased.connect(self.ccd1expose)
        self.horizontalSlider_2.sliderReleased.connect(self.ccd2expose)
        self.horizontalSlider_3.sliderReleased.connect(self.ccd3expose)
        self.pushButton_16.clicked.connect(self.ccd1)  # 相机1图像开启
        self.pushButton_17.clicked.connect(self.ccd2)  # 相机2图像开启
        self.pushButton_18.clicked.connect(self.ccd3)  # 相机3图像开启
        self.pushButton_14.clicked.connect(self.ccd1_roi1)  # 相机1图像roi1获取
        self.pushButton_15.clicked.connect(self.ccd1_roi2)  # 相机1图像roi2获取
        self.pushButton_25.clicked.connect(self.ccd1_roi1p)# 相机1图像roi1p获取
        self.pushButton_9.clicked.connect(self.ccd2_roi1)  # 相机2图像roi1获取
        self.pushButton_11.clicked.connect(self.ccd2_roi2)  # 相机2图像roi2获取
        self.pushButton_12.clicked.connect(self.ccd3_roi1)  # 相机3图像roi1获取
        self.pushButton_13.clicked.connect(self.ccd3_roi2)  # 相机3图像roi2获取
        self.pushButton_19.clicked.connect(self.ccd1roi1yuzhi)  # 相机1图像roi1阈值获取
        self.pushButton_20.clicked.connect(self.ccd1roi2yuzhi)  # 相机1图像roi2阈值获取
        self.pushButton_22.clicked.connect(self.ccd2roi1yuzhi)  # 相机2图像roi1阈值获取
        self.pushButton_21.clicked.connect(self.ccd2roi2yuzhi)  # 相机2图像roi2阈值获取
        self.pushButton_23.clicked.connect(self.ccd3roi1yuzhi)  # 相机3图像roi1阈值获取
        self.pushButton_24.clicked.connect(self.ccd3roi2yuzhi)  # 相机3图像roi2阈值获取
    def ccd1roi1yuzhi(self):
        print("阈值确认")
        self.ccdyuzhi[0]=yuzhi
        print(self.ccdyuzhi[0])
    def ccd1roi2yuzhi(self):
        print("阈值确认")
        self.ccdyuzhi[1]=yuzhi
        print(self.ccdyuzhi[1])
    def ccd2roi1yuzhi(self):
        print("阈值确认")
        self.ccd2yuzhi[0]=yuzhi
    def ccd2roi2yuzhi(self):
        print("阈值确认")
        self.ccd2yuzhi[1]=yuzhi
    def ccd3roi1yuzhi(self):
        print("阈值确认")
        self.ccd3yuzhi[0]=yuzhi
    def ccd3roi2yuzhi(self):
        print("阈值确认")
        self.ccd3yuzhi[1]=yuzhi
    #######相机一roi获取
    def ccd1_roi1(self):
        global point1, point2
        md.ccd1_roi1x = point1[1]  # y  转到像素坐标
        md.ccd1_roi1y = point1[0]  # x   转到像素坐标
        print(md.ccd1_roi1x, md.ccd1_roi1y)
        md.ccd1_roi1xend = point2[1]  # y  转到像素坐标
        md.ccd1_roi1yend = point2[0]  # x   转到像素坐标
    def ccd1_roi1p(self):
        global point1, point2
        md.ccd1_roi1xp = point1[1]  # y  转到像素坐标
        md.ccd1_roi1yp = point1[0]  # x   转到像素坐标
        print(md.ccd1_roi1xp, md.ccd1_roi1yp)
        md.ccd1_roi1xendp = point2[1]  # y  转到像素坐标
        md.ccd1_roi1yendp = point2[0]  # x   转到像素坐标
    def ccd1_roi2(self):
        global point1,point2
        md.ccd1_roi2x = point1[1]  # y  转到像素坐标
        md.ccd1_roi2y = point1[0]  # x   转到像素坐标
        md.ccd1_roi2xend = point2[1]  # y  转到像素坐标
        md.ccd1_roi2yend = point2[0]  # x   转到像素坐标
        print(md.ccd1_roi2x, md.ccd1_roi2y)

    #################相机二roi获取
    def ccd2_roi1(self):
        global point1,point2
        md.ccd2_roi1x = point1[1]  # y  转到像素坐标
        md.ccd2_roi1y = point1[0]  # x   转到像素坐标
        md.ccd2_roi1xend = point2[1]
        md.ccd2_roi1yend = point2[0]
    def ccd2_roi2(self):
        global point1,point2
        md.ccd2_roi2x = point1[1]  # y  转到像素坐标
        md.ccd2_roi2y = point1[0]  # x   转到像素坐标
        md.ccd2_roi2xend = point2[1]  # y  转到像素坐标
        md.ccd2_roi2yend = point2[0]  # x   转到像素坐标

    #################相机三roi获取
    def ccd3_roi1(self):
        global point1,point2
        md.ccd3_roi1x = point1[1]  # y  转到像素坐标
        md.ccd3_roi1y = point1[0]  # x   转到像素坐标
        md.ccd3_roi1xend = point2[1]  # y  转到像素坐标
        md.ccd3_roi1yend = point2[0]  # x   转到像素坐标
    def ccd3_roi2(self):
        global point1,point2
        md.ccd3_roi2x = point1[1]  # y  转到像素坐标
        md.ccd3_roi2y = point1[0]  # x   转到像素坐标
        md.ccd3_roi2xend = point2[1]  # y  转到像素坐标
        md.ccd3_roi2yend = point2[0]  # x   转到像素坐标

    def ccd1(self):
        
        md.cam.ExposureTime.set(md.ccd1_exposure_time)
        # set gain
        md.cam.Gain.set(10.0)
        # 流开启
        md.cam.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1

        for i in range(num):
            # get raw image
            raw_image = md.cam.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam.stream_off()
        
        create_roi_automatically(numpy_image) #ccd1的阈值

    def ccd2(self):
        md.cam2.ExposureTime.set(md.ccd2_exposure_time)
        # set gain
        md.cam2.Gain.set(10.0)
        # 流开启
        md.cam2.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam2.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = md.cam2.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam2.stream_off()
        create_roi_automatically(numpy_image,1)

    def ccd3(self):
        md.cam3.ExposureTime.set(md.ccd3_exposure_time)
        # set gain
        md.cam3.Gain.set(10.0)
        # 流开启
        md.cam3.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam3.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = md.cam3.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam3.stream_off()
        create_roi_automatically(numpy_image, 2)

    def ccd1expose(self):
        md.ccd1_exposure_time = self.horizontalSlider.value()
        md.sheet1['B12'] = md.ccd1_exposure_time  # ccd1_exposure_time
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")

    def ccd2expose(self):
        md.ccd2_exposure_time = self.horizontalSlider_2.value()
        md.sheet1['B13'] = md.ccd2_exposure_time  # ccd2_exposure_time
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")

    def ccd3expose(self):
        md.ccd3_exposure_time = self.horizontalSlider_3.value()
        md.sheet1['B14'] = md.ccd3_exposure_time  # ccd3_exposure_time
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")

    def load_parameter(self):
        # 加载文档
        self.lineEdit_9.setText(str(md.sheet1.cell(row=2, column=2).value))  # cam1 K
        self.lineEdit_10.setText(str(md.sheet1.cell(row=3, column=2).value))  # cam1 b
        self.lineEdit_5.setText(str(md.sheet1.cell(row=4, column=2).value))  # cam1 补偿
        self.lineEdit_11.setText(str(md.sheet1.cell(row=5, column=2).value))  # cam2 k
        self.lineEdit_12.setText(str(md.sheet1.cell(row=6, column=2).value))  # cam2 b
        self.lineEdit_6.setText(str(md.sheet1.cell(row=7, column=2).value))  # cam2 补偿
        self.lineEdit_13.setText(str(md.sheet1.cell(row=8, column=2).value))  # cam3 k
        self.lineEdit_14.setText(str(md.sheet1.cell(row=9, column=2).value))  # cam3  b
        self.lineEdit_7.setText(str(md.sheet1.cell(row=10, column=2).value))  # cam3  补偿
        #加载CCD阈值
        self.ccdyuzhi[0]= md.sheet1.cell(row=33, column=2).value
        self.ccdyuzhi[1] = md.sheet1.cell(row=34, column=2).value
        self.ccd2yuzhi[0] = md.sheet1.cell(row=35, column=2).value
        self.ccd2yuzhi[1] = md.sheet1.cell(row=36, column=2).value
        self.ccd3yuzhi[0] = md.sheet1.cell(row=37, column=2).value
        self.ccd3yuzhi[1] = md.sheet1.cell(row=38, column=2).value
        print("load done")

    def save_parameter(self):
        print("save parameter")
        md.sheet1['B2'] = float(self.lineEdit_9.text())  # cam1 K
        md.sheet1['B3'] = float(self.lineEdit_10.text())  # cam1 b
        md.sheet1['B4'] = float(self.lineEdit_5.text())  # cam1 补偿
        md.sheet1['B5'] = float(self.lineEdit_11.text())  # cam2 k
        md.sheet1['B6'] = float(self.lineEdit_12.text())  # cam2 b
        md.sheet1['B7'] = float(self.lineEdit_6.text())  # cam2 补偿
        md.sheet1['B8'] = float(self.lineEdit_13.text())  # cam3 k
        md.sheet1['B9'] = float(self.lineEdit_14.text())  # cam3  b
        md.sheet1['B10'] = float(self.lineEdit_7.text())  # cam3  补偿
        md.sheet1['B15'] = md.ccd1_roi1x
        md.sheet1['B16'] = md.ccd1_roi1y
        md.sheet1['B17'] = md.ccd1_roi2x
        md.sheet1['B18'] = md.ccd1_roi2y
        md.sheet1['B19'] = md.ccd2_roi1x
        md.sheet1['B20'] = md.ccd2_roi1y
        md.sheet1['B21'] = md.ccd2_roi2x
        md.sheet1['B22'] = md.ccd2_roi2y
        md.sheet1['B23'] = md.ccd3_roi1x
        md.sheet1['B24'] = md.ccd3_roi1y
        md.sheet1['B25'] = md.ccd3_roi2x
        md.sheet1['B26'] = md.ccd3_roi2y
        md.sheet1['B33'] = self.ccdyuzhi[0]
        md.sheet1['B34'] = self.ccdyuzhi[1]
        md.sheet1['B35'] = self.ccd2yuzhi[0]
        md.sheet1['B36'] = self.ccd2yuzhi[1]
        md.sheet1['B37'] = self.ccd3yuzhi[0]
        md.sheet1['B38'] = self.ccd3yuzhi[1]
        md.sheet1['B39'] = md.ccd1_roi1xend
        md.sheet1['B40'] = md.ccd1_roi1yend
        md.sheet1['B41'] = md.ccd1_roi2xend
        md.sheet1['B42'] = md.ccd1_roi2yend
        md.sheet1['B43'] = md.ccd2_roi1xend
        md.sheet1['B44'] = md.ccd2_roi1yend
        md.sheet1['B45'] = md.ccd2_roi2xend
        md.sheet1['B46'] = md.ccd2_roi2yend
        md.sheet1['B47'] = md.ccd3_roi1xend
        md.sheet1['B48'] = md.ccd3_roi1yend
        md.sheet1['B49'] = md.ccd3_roi2xend
        md.sheet1['B50'] = md.ccd3_roi2yend
        md.sheet1['B52'] = md.ccd1_roi1xp
        md.sheet1['B53'] = md.ccd1_roi1yp
        md.sheet1['B54'] = md.ccd1_roi1xendp
        md.sheet1['B55'] = md.ccd1_roi1yendp
        print("parameter get")
        md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
        print("saved done")

    def canshuqueren(self):
        print("参数确认状态")
        global camera1k, camera1b
        global camera2k, camera2b
        global camera3k, camera3b
        camera1k, camera1b = float(self.lineEdit_9.text()), float(self.lineEdit_10.text())
        camera2k, camera2b = float(self.lineEdit_11.text()), float(self.lineEdit_12.text())
        camera3k, camera3b = float(self.lineEdit_13.text()), float(self.lineEdit_14.text())
        print("get factors ")
        self.label_17.setText("Done")

    def zichuangkouclose(self):
        self.close()

    def camera_cali(self):
        global camera1k, camera1b,camera2k
        global shuzu
        shuzu = list()
        
        md.cam.ExposureTime.set(md.ccd1_exposure_time)
        # set gain
        md.cam.Gain.set(10.0)
        # 流开启
        md.cam.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = md.cam.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam.stream_off()
        
        # img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # img = cv2.resize(numpy_image, (int(numpy_image.shape[1] / 3), int(numpy_image.shape[0] / 3)))
        if md.camera1_strategy == int(0):  # 脱水轴（大油封）高度
            distance1, img = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=md.ccd1_roi1x,
                                                                           roi1y=md.ccd1_roi1y,
                                                                           roi1xend=md.ccd1_roi1xend,
                                                                           roi1yend=md.ccd1_roi1yend,
                                                                           roi2x=md.ccd1_roi2x, roi2y=md.ccd1_roi2y,
                                                                           roi2xend=md.ccd1_roi2xend,
                                                                           roi2yend=md.ccd1_roi2yend,
                                                                           thresh0=self.ccdyuzhi[0],
                                                                           thresh1=self.ccdyuzhi[
                                                                               1])  # 长度距离检测
            if md.label_window.liheqi==int(0):
                distance2, img2 = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=md.ccd1_roi1xp,
                                                                           roi1y=md.ccd1_roi1yp,
                                                                           roi1xend=md.ccd1_roi1xendp,
                                                                           roi1yend=md.ccd1_roi1yendp,
                                                                           roi2x=md.ccd1_roi2x, roi2y=md.ccd1_roi2y,
                                                                           roi2xend=md.ccd1_roi2xend,
                                                                           roi2yend=md.ccd1_roi2yend,
                                                                           thresh0=self.ccdyuzhi[0],
                                                                           thresh1=self.ccdyuzhi[
                                                                               1])  # 长度距离检测
                cv2.imshow("image", img2)
                cv2.waitKey(0)
                cv2.destroyAllWindows()
                # 单个像素值实际距离计算公式：k=实际距离/像素点长度
                camera2k = float(self.lineEdit_2.text()) / distance2
                print(camera2k)
                self.label_9.setText("camera2k is" + str(camera2k))
        elif md.camera1_strategy == int(1):  # 制动臂距离
            # 图像处理
            distance1, img = procCallback.cepianyizhi(numpy_image, roi1x=md.ccd1_roi1x,
                                                               roi1y=md.ccd1_roi1y, roi1xend=self.ccd1_roi1xend,
                                                               roi1yend=md.ccd1_roi1yend,
                                                               roi2x=md.ccd1_roi2x, roi2y=self.ccd1_roi2y,
                                                               roi2xend=md.ccd1_roi2xend,
                                                               roi2yend=md.ccd1_roi2yend,
                                                               thresh0=self.ccdyuzhi[0],
                                                               thresh1=self.ccdyuzhi[1])  # 偏心距
        elif md.camera1_strategy == int(2):  # 间隙
            # 图像处理
            distance1, img = procCallback.measure_gear_gap(numpy_image, roi1x=md.ccd1_roi1x,
                                                                    roi1y=md.ccd1_roi1y,
                                                                    roi1xend=md.ccd1_roi1xend,
                                                                    roi1yend=md.ccd1_roi1yend,
                                                                    roi2x=md.ccd1_roi2x, roi2y=self.ccd1_roi2y,
                                                                    roi2xend=md.ccd1_roi2xend,
                                                                    roi2yend=md.ccd1_roi2yend,
                                                                    thresh0=self.ccdyuzhi[0],
                                                                    thresh1=self.ccdyuzhi[1])
       # distance1, img = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=md.ccd1_roi1x, roi1y=md.ccd1_roi1y,roi1xend=md.ccd1_roi1xend,roi1yend=md.ccd1_roi1yend,
          #                                                         roi2x = md.ccd1_roi2x, roi2y=md.ccd1_roi2y,roi2xend=md.ccd1_roi2xend,roi2yend=md.ccd1_roi2yend,
          #                                                         thresh0 =self.ccdyuzhi[0],thresh1 = self.ccdyuzhi[1])
        # 单个像素值实际距离计算公式：k=实际距离/像素点长度
        cv2.imshow("image", img)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        camera1k = float(self.lineEdit.text()) / distance1#标定比例系数
        print(camera1k)
        self.label_8.setText("camera1k is" + str(camera1k))

    def camera_cali2(self):
        global shuzu
        global camera2k, camera2b
        shuzu = list()
        md.cam2.ExposureTime.set(md.ccd2_exposure_time)
        # set gain
        md.cam2.Gain.set(10.0)
        # 流开启
        md.cam2.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam2.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = md.cam2.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam2.stream_off()
        # img = cv2.imread('2022-05-26_20_34_01_975.bmp')
        # img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # img = cv2.resize(numpy_image, (int(numpy_image.shape[1] / 3), int(numpy_image.shape[0] / 3)))
        if not md.label_window.liheqi==int(0):
            if md.camera2_strategy == int(0):  # 脱水轴（大油封）高度
                distance2, img = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=md.ccd2_roi1x,
                                                                               roi1y=md.ccd2_roi1y,
                                                                               roi1xend=md.ccd2_roi1xend,
                                                                               roi1yend=md.ccd2_roi1yend,
                                                                               roi2x=md.ccd2_roi2x, roi2y=md.ccd2_roi2y,
                                                                               roi2xend=md.ccd2_roi2xend,
                                                                               roi2yend=md.ccd2_roi2yend,
                                                                               thresh0=self.ccd2yuzhi[0],
                                                                               thresh1=self.ccd2yuzhi[
                                                                                   1])  # 长度距离检测


            elif md.camera2_strategy == int(1):  # 制动臂距离

                # 图像处理
                distance2, img = procCallback.cepianyizhi(numpy_image, roi1x=md.ccd2_roi1x,
                                                                   roi1y=md.ccd2_roi1y, roi1xend=md.ccd2_roi1xend,
                                                                   roi1yend=md.ccd2_roi1yend,
                                                                   roi2x=md.ccd2_roi2x, roi2y=md.ccd2_roi2y,
                                                                   roi2xend=md.ccd2_roi2xend,
                                                                   roi2yend=md.ccd2_roi2yend,
                                                                   thresh0=self.ccd2yuzhi[0],
                                                                   thresh1=self.ccd2yuzhi[1])  # 偏心距
                print("run here")
            elif md.camera2_strategy == int(2):  # 间隙
                # 图像处理
                distance2, img = procCallback.measure_gear_gap(numpy_image, roi1x=md.ccd2_roi1x,
                                                                        roi1y=md.ccd2_roi1y,
                                                                        roi1xend=md.ccd2_roi1xend,
                                                                        roi1yend=md.ccd2_roi1yend,
                                                                        roi2x=md.ccd2_roi2x, roi2y=md.ccd2_roi2y,
                                                                        roi2xend=md.ccd2_roi2xend,
                                                                        roi2yend=md.ccd2_roi2yend,
                                                                        thresh0=self.ccd2yuzhi[0],
                                                                        thresh1=self.ccd2yuzhi[1])
            #distance2, img = procCallback.cepianyizhi(numpy_image, roi1x=md.ccd2_roi1x, roi1y=md.ccd2_roi1y,roi1xend=md.ccd2_roi1xend,roi1yend=md.ccd2_roi1yend,
            #                                                           roi2x = md.ccd2_roi2x, roi2y=md.ccd2_roi2y,roi2xend=md.ccd2_roi2xend,roi2yend=md.ccd2_roi2yend,
             #                                                          thresh0 =self.ccd2yuzhi[0],thresh1 = self.ccd2yuzhi[1])

            cv2.imshow("image", img)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
            # 单个像素值实际距离计算公式：k=实际距离/像素点长度
            camera2k = float(self.lineEdit_2.text()) / distance2
            print(camera2k)
            self.label_9.setText("camera2k is" + str(camera2k))

    def camera_cali3(self):
        global camera3k, camera3b
        global shuzu
        shuzu = list()
        md.cam3.ExposureTime.set(md.ccd3_exposure_time)
        # set gain
        md.cam3.Gain.set(10.0)
        # 流开启
        md.cam3.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        md.cam3.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = md.cam3.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
        md.cam3.stream_off()

        # img = cv2.imread('2022-05-26_20_34_01_975.bmp')
        # img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # img = cv2.resize(numpy_image, (int(numpy_image.shape[1] / 3), int(numpy_image.shape[0] / 3)))
        # img = cv2.resize(numpy_image, (int(numpy_image.shape[1] / 3), int(numpy_image.shape[0] / 3)))
        if md.camera3_strategy == int(0):  # 脱水轴（大油封）高度
            distance3, img = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=md.ccd3_roi1x,
                                                                           roi1y=md.ccd3_roi1y,
                                                                           roi1xend=md.ccd3_roi1xend,
                                                                           roi1yend=md.ccd3_roi1yend,
                                                                           roi2x=md.ccd3_roi2x, roi2y=md.ccd3_roi2y,
                                                                           roi2xend=md.ccd3_roi2xend,
                                                                           roi2yend=md.ccd3_roi2yend,
                                                                           thresh0=self.ccd3yuzhi[0],
                                                                           thresh1=self.ccd3yuzhi[
                                                                               1])  # 长度距离检测


        elif md.camera3_strategy == int(1):  # 制动臂距离
            # 图像处理
            distance3, img = procCallback.cepianyizhi(numpy_image, roi1x=md.ccd3_roi1x,
                                                               roi1y=md.ccd3_roi1y, roi1xend=md.ccd3_roi1xend,
                                                               roi1yend=md.ccd3_roi1yend,
                                                               roi2x=md.ccd3_roi2x, roi2y=md.ccd3_roi2y,
                                                               roi2xend=md.ccd3_roi2xend,
                                                               roi2yend=md.ccd3_roi2yend,
                                                               thresh0=self.ccd3yuzhi[0],
                                                               thresh1=self.ccd3yuzhi[1])  # 偏心距
        elif md.camera3_strategy == int(2):  # 间隙
            # 图像处理
            distance3, img = procCallback.measure_gear_gap(numpy_image, roi1x=md.ccd3_roi1x,
                                                                    roi1y=md.ccd3_roi1y,
                                                                    roi1xend=md.ccd3_roi1xend,
                                                                    roi1yend=md.ccd3_roi1yend,
                                                                    roi2x=md.ccd3_roi2x, roi2y=md.ccd3_roi2y,
                                                                    roi2xend=md.ccd3_roi2xend,
                                                                    roi2yend=md.ccd3_roi2yend,
                                                                    thresh0=self.ccd3yuzhi[0],
                                                                    thresh1=self.ccd3yuzhi[1])
        #distance3, img = procCallback.measure_gear_gap(numpy_image, roi1x=md.ccd3_roi1x, roi1y=md.ccd3_roi1y,roi1xend=md.ccd3_roi1xend,roi1yend=md.ccd3_roi1yend,
        #                                                           roi2x = md.ccd3_roi2x, roi2y=md.ccd3_roi2y,roi2xend=md.ccd3_roi2xend,roi2yend=md.ccd3_roi2yend,
         #                                                          thresh0 =self.ccd3yuzhi[0],thresh1 = self.ccd3yuzhi[1])
        cv2.imshow("image", img)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        # 单个像素值实际距离计算公式：k=实际距离/像素点长度
        camera3k = float(self.lineEdit_3.text()) / distance3
        print(camera3k)
        self.label_10.setText("camera3k is" + str(camera3k))


##############################主窗口类create class for UI#############################
class MainCode(QMainWindow, sanxing08.Ui_mainWindow):
    def __init__(self):
        QMainWindow.__init__(self)
        sanxing08.Ui_mainWindow.__init__(self)
        self.setupUi(self)
        self.initial()
        self.painter = QPainter(self)
        self.actioncamerabiaoding.triggered.connect(self.zichuangkou1)
        self.actionbiaoqianxinxi.triggered.connect(self.labelinfotrig)
        self.camera1flag = False
        self.camera2flag = False
        self.camera3flag = False
        self.print_sensor = threading.Thread(target=self.print_trig)
        self.print_sensor.setDaemon(True)
        self.print_sensor.start()
        self.comboBox.currentIndexChanged.connect(self.selectionchange)
        self.comboBox_2.currentIndexChanged.connect(self.selectionchange_2)
        self.comboBox_3.currentIndexChanged.connect(self.selectionchange_3)
    def initial_para(self):
        self.fonts = 'C:/Windows/Fonts/ARLRDBD.TTF'
        # fnt3 = ImageFont.truetype('C:/Windows/Fonts/STXIHEI.TTF', 50)  # 细黑
        self.workbook_initial = openpyxl.load_workbook('parameter_record.xlsx')  # 返回一个workbook数据类型的值.
        self.sheet1 = self.workbook_initial.active  # 获取活动表
        data = self.sheet1.cell(row=1, column=2)
        zubie = self.sheet1.cell(row=11, column=2)
        scxian = self.sheet1.cell(row=56, column=2)
        ccd1_expose_time = self.sheet1.cell(row=12, column=2)
        ccd2_expose_time = self.sheet1.cell(row=13, column=2)
        ccd3_expose_time = self.sheet1.cell(row=14, column=2)
        ccd1_roi1x = self.sheet1.cell(row=15, column=2)
        ccd1_roi1y = self.sheet1.cell(row=16, column=2)
        ccd1_roi1xp = self.sheet1.cell(row=52, column=2)
        ccd1_roi1yp = self.sheet1.cell(row=53, column=2)
        ccd1_roi2x = self.sheet1.cell(row=17, column=2)
        ccd1_roi2y = self.sheet1.cell(row=18, column=2)
        ccd2_roi1x = self.sheet1.cell(row=19, column=2)
        ccd2_roi1y = self.sheet1.cell(row=20, column=2)
        ccd2_roi2x = self.sheet1.cell(row=21, column=2)
        ccd2_roi2y = self.sheet1.cell(row=22, column=2)
        ccd3_roi1x = self.sheet1.cell(row=23, column=2)
        ccd3_roi1y = self.sheet1.cell(row=24, column=2)
        ccd3_roi2x = self.sheet1.cell(row=25, column=2)
        ccd3_roi2y = self.sheet1.cell(row=26, column=2)
        # print(data.value)
        # ------------------------------------------------------------#
        self.serial_number = data.value
        self.zubie = zubie.value
        self.scxian = scxian.value
        self.ccd1_exposure_time = ccd1_expose_time.value
        self.ccd2_exposure_time = ccd2_expose_time.value
        self.ccd3_exposure_time = ccd3_expose_time.value
        self.ccd1_roi1x = ccd1_roi1x.value
        self.ccd1_roi1y = ccd1_roi1y.value
        self.ccd1_roi1xp = ccd1_roi1xp.value
        self.ccd1_roi1yp = ccd1_roi1yp.value
        self.ccd1_roi2x = ccd1_roi2x.value
        self.ccd1_roi2y = ccd1_roi2y.value
        self.ccd2_roi1x = ccd2_roi1x.value
        self.ccd2_roi1y = ccd2_roi1y.value
        self.ccd2_roi2x = ccd2_roi2x.value
        self.ccd2_roi2y = ccd2_roi2y.value
        self.ccd3_roi1x = ccd3_roi1x.value
        self.ccd3_roi1y = ccd3_roi1y.value
        self.ccd3_roi2x = ccd3_roi2x.value
        self.ccd3_roi2y = ccd3_roi2y.value
        self.ccd1_roi1xend = self.sheet1.cell(row=39, column=2).value
        self.ccd1_roi1yend = self.sheet1.cell(row=40, column=2).value
        self.ccd1_roi1xendp = self.sheet1.cell(row=54, column=2).value
        self.ccd1_roi1yendp = self.sheet1.cell(row=55, column=2).value
        self.ccd1_roi2xend = self.sheet1.cell(row=41, column=2).value
        self.ccd1_roi2yend = self.sheet1.cell(row=42, column=2).value
        self.ccd2_roi1xend = self.sheet1.cell(row=43, column=2).value
        self.ccd2_roi1yend = self.sheet1.cell(row=44, column=2).value
        self.ccd2_roi2xend = self.sheet1.cell(row=45, column=2).value
        self.ccd2_roi2yend = self.sheet1.cell(row=46, column=2).value
        self.ccd3_roi1xend = self.sheet1.cell(row=47, column=2).value
        self.ccd3_roi1yend = self.sheet1.cell(row=48, column=2).value
        self.ccd3_roi2xend = self.sheet1.cell(row=49, column=2).value
        self.ccd3_roi2yend = self.sheet1.cell(row=50, column=2).value
        self.label_window.liheqi = self.sheet1.cell(row=51, column=2).value#加载离合器选型参数
        print("label_window.liheqi is",self.label_window.liheqi)
        self.label_window.lineEdit_5.setText(self.scxian)
        self.label_window.lineEdit_45.setText(self.zubie)
    def initial(self):
        self.label_window = labelwindow()  # 绑定标签窗口到主窗口
        self.chile_window = biaodingwindow()  # 绑定相机标定窗口到主窗口
        self.initial_para()
        self.cam1result = 0
        self.cam2result = 0
        self.cam3result = 0
        self.label_33.setText(" ")
        self.label_10.setPixmap(QPixmap("blank01.jpg", ))
        self.label_11.setPixmap(QPixmap("blank01.jpg", ))
        self.label_12.setPixmap(QPixmap("blank01.jpg", ))
        self.L = 60  # 圆心到杆的距离39毫米
        self.width_zhidongbi = 11.5  # 制动臂宽度
        # self.leftL=49.5 #制动臂左侧到杆的距离
        self.camera1_strategy = int(0)
        self.camera2_strategy = int(1)
        self.camera3_strategy = int(2)
        self.triggerflag = False
        self.num = 0  # 总数
        self.cam1good = 0
        self.cam1goodrate = 100
        self.cam2good = 0
        self.cam2goodrate = 100
        self.cam3good = 0
        self.cam3goodrate = 100
        self.label_14.setText(str(self.num))
        # self.label_24.setText(str(self.cam1good))
        # self.label_22.setText(str(self.cam1goodrate))
        # self.label_25.setText(str(self.cam2good))
        # self.label_23.setText(str(self.cam2goodrate))
        # self.label_21.setText(str(self.cam3good))
        # self.label_26.setText(str(self.cam3goodrate))
        self.triggerflag = False
        print("Initializing camera......")
        # create a device manager
        self.device_manager = gx.DeviceManager()
        dev_num, dev_info_list = self.device_manager.update_device_list()
        print(dev_info_list)
        if dev_num is 0:
            print("Number of enumerated devices is 0")
            return
        else:
            print("get device")
        # open the first device
        self.cam = self.device_manager.open_device_by_index(2)
        self.cam2 = self.device_manager.open_device_by_index(1)
        self.cam3 = self.device_manager.open_device_by_index(3)
        self.cam.TriggerMode.set(gx.GxSwitchEntry.ON)  # 设置触发条件 开启
        self.cam2.TriggerMode.set(gx.GxSwitchEntry.ON)  # 设置触发条件 开启
        self.cam3.TriggerMode.set(gx.GxSwitchEntry.ON)  # 设置触发条件 开启
        # 监听line0当前状态
        self.cam.LineSelector.set(0)
        print("status is :", self.cam.LineStatus.get())
        self.sensortrig = threading.Thread(target=self.trigersensor)
        self.sensortrig.setDaemon(True)
        self.sensortrig.start()

    def labelinfotrig(self):  # 进入标签信息设置界面
        print("进入标签")
        self.label_window.show()

    def zichuangkou1(self):  # 进入相机标定子界面
        # self.camera_califlag=True
        print("进入标定")
        self.chile_window.show()

    def trigersensor(self):
        # self.triggerflag = False
        last_state = False
        self.cam.LineSelector.set(0)
        # print("status is :", self.cam.LineStatus.get())
        while 1:
            # 监听line0当前状态
            # print("current line2 status is :", self.cam.LineStatus.get())
            # self.cam.LineSelector.set(0)
            current_state = self.cam.LineStatus.get()
            if last_state == 0 and current_state == 1:  # 上升沿触发
                self.starttime = str(time.strftime('%Y%m%d%H%M%S', time.localtime(time.time())))
                self.triggerflag = True
                self.num = self.num + 1
                self.label_14.setText(str(self.num))
                self.camera1()
                self.camera2()
                self.camera3()
            else:
                # print("i am here,no photo took")
                pass
            last_state = current_state

    def camera1(self):
        global camera1k, camera1b
        self.cam.ExposureTime.set(self.ccd1_exposure_time)
        # set gain
        self.cam.Gain.set(10.0)
        # 流开启
        self.cam.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        self.cam.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = self.cam.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
            # todo   检测策略的选择
            # 检测策略的选择
            # 图像处理
            if self.camera1_strategy == int(0): #脱水轴高度
                self.gap, self.img = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=self.ccd1_roi1x, roi1y=self.ccd1_roi1y,roi1xend=self.ccd1_roi1xend,roi1yend=self.ccd1_roi1yend,
                                                                   roi2x = self.ccd1_roi2x, roi2y=self.ccd1_roi2y,roi2xend=self.ccd1_roi2xend,roi2yend=self.ccd1_roi2yend,
                                                                   thresh0 =self.chile_window.ccdyuzhi[0],thresh1 = self.chile_window.ccdyuzhi[1]) # 长度距离检测
                print("distance is :", self.gap)
                self.gap = round(camera1k, 15) * self.gap
                self.cam1result = self.gap
                if not self.label_window.liheqi==int(2):
                    self.gap2, self.img2 = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=self.ccd1_roi1xp,
                                                                                   roi1y=self.ccd1_roi1yp,
                                                                                   roi1xend=self.ccd1_roi1xendp,
                                                                                   roi1yend=self.ccd1_roi1yendp,
                                                                                   roi2x=self.ccd1_roi2x,
                                                                                   roi2y=self.ccd1_roi2y,
                                                                                   roi2xend=self.ccd1_roi2xend,
                                                                                   roi2yend=self.ccd1_roi2yend,
                                                                                   thresh0=self.chile_window.ccdyuzhi[
                                                                                       0],
                                                                                   thresh1=self.chile_window.ccdyuzhi[
                                                                                       1])  # 长度距离检测
                    print("distance is :", self.gap2)
                    self.cam2result = self.gap2
                    self.distance = round(camera2k, 15) * self.gap2
                    self.cam2result = self.distance


            elif self.camera1_strategy == int(1):#制动臂距离
                # 图像处理
                self.distance, self.img = procCallback.cepianyizhi(numpy_image, roi1x=self.ccd1_roi1x,
                                                                    roi1y=self.ccd1_roi1y, roi1xend=self.ccd1_roi1xend,
                                                                    roi1yend=self.ccd1_roi1yend,
                                                                    roi2x=self.ccd1_roi2x, roi2y=self.ccd1_roi2y,
                                                                    roi2xend=self.ccd1_roi2xend,
                                                                    roi2yend=self.ccd1_roi2yend,
                                                                    thresh0=self.chile_window.ccdyuzhi[0],
                                                                    thresh1=self.chile_window.ccdyuzhi[1])  # 偏心距
                print("虚拟pianxiju:", self.distance)
                self.distance = round(camera1k, 15) * self.distance
                # todo distance为轴到标定板的距离
                # 实际偏心距离distance = distance - 实际中心距离L
                # self.distance = self.distance - self.L
                # 制动臂左侧到圆心距离
                self.cam1result = self.distance
            elif self.camera1_strategy == int(2):#间隙
                # 图像处理
                self.distance, self.img = procCallback.measure_gear_gap(numpy_image, roi1x=self.ccd1_roi1x,
                                                                         roi1y=self.ccd1_roi1y,
                                                                         roi1xend=self.ccd1_roi1xend,
                                                                         roi1yend=self.ccd1_roi1yend,
                                                                         roi2x=self.ccd1_roi2x, roi2y=self.ccd1_roi2y,
                                                                         roi2xend=self.ccd1_roi2xend,
                                                                         roi2yend=self.ccd1_roi2yend,
                                                                         thresh0=self.chile_window.ccd1yuzhi[0],
                                                                         thresh1=self.chile_window.ccd1yuzhi[1])
                print("distance3:", self.distance)
                self.distance = round(camera1k, 15) * self.distance
                # 制动臂左侧到圆心距离
                self.cam1result = self.distance
            # todo 误差补偿
            self.cam1result = round((self.cam1result + (float(self.chile_window.lineEdit_5.text()))), 2)
            print('cam1result', self.cam1result)
            self.label_30.setText("相机一结果" + str(self.cam1result))
            if self.label_window.liheqi == int(0):
                # todo 结果检查
                if self.cam1result <= float(self.label_window.lineEdit_20.text()) and self.cam1result >= float(
                        self.label_window.lineEdit_19.text()):  # 检测条件构建
                    self.cam1good = self.cam1good + 1
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam1good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B27'] = self.cam1good
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera1flag = True
                    self.label_10.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_10.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera1flag = False
            elif self.label_window.liheqi == int(1):
                # todo 结果检查
                if self.cam1result <= float(self.label_window.lineEdit_42.text()) and self.cam1result >= float(
                        self.label_window.lineEdit_31.text()):  # 检测条件构建
                    self.cam1good = self.cam1good + 1
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam1good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B27'] = self.cam1good
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera1flag = True
                    self.label_10.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_10.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera1flag = False
            elif self.label_window.liheqi==int(2):
            # todo 结果检查
                if self.cam1result <=float(self.label_window.lineEdit_42.text()) and self.cam1result >= float(
                        self.label_window.lineEdit_31.text()):  # 检测条件构建
                    self.cam1good = self.cam1good + 1
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam1good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B27'] = self.cam1good
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera1flag = True
                    self.label_10.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam1goodrate = self.cam1good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_22.setText(str(self.cam1goodrate))
                    self.sheet1['B28'] = self.cam1goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_10.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera1flag = False
            # show acquired image
            # img = Image.fromarray(self.img, 'L')
            # print("imgformate is:", type(img))
            # self.cv2img = cv2.cvtColor(np.asarray(img), cv2.COLOR_GRAY2RGB)
            # self.cv2img = cv2.resize(self.cv2img, (450, 450))
            self.cv2img = cv2.resize(self.img, (450, 450))
            # cv2.imshow("this is cv2 window", cv2img)
            # cv2.waitKey(0)
            self.Qframe_daheng1 = QImage(self.cv2img.data, self.cv2img.shape[1], self.cv2img.shape[0],
                                         self.cv2img.shape[1] * 3,
                                         QImage.Format_RGB888)
            # print(Qframe)
            # pix = QPixmap(Qframe).scaled(frame.shape[1], frame.shape[0])
            # self.setPixmap(pix)
            # QRect qq(20,50,self.img.width,self.img.height)
            self.label.setPixmap(QPixmap.fromImage(self.Qframe_daheng1))
            print("show success")
            # print height, width, and frame ID of the acquisition image
            print("Frame ID: %d   Height: %d   Width: %d"
                  % (raw_image.get_frame_id(), raw_image.get_height(), raw_image.get_width()))
        # self.cam.stream_off()
        # self.cam.close_device()

    def camera2(self):
        global camera2k, camera2b

        self.cam2.ExposureTime.set(self.ccd2_exposure_time)
        # set gain
        self.cam2.Gain.set(10.0)
        # 流开启
        self.cam2.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        self.cam2.TriggerSoftware.send_command()  # 软件触发
        # time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = self.cam2.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue

            # 检测策略的选择
            # 图像处理相机2
            if self.label_window.liheqi==int(2):
                if self.camera2_strategy == int(0):  # 脱水轴高度
                    self.gap, self.img2 = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=self.ccd2_roi1x,
                                                                                   roi1y=self.ccd2_roi1y,
                                                                                   roi1xend=self.ccd2_roi1xend,
                                                                                   roi1yend=self.ccd2_roi1yend,
                                                                                   roi2x=self.ccd2_roi2x,
                                                                                   roi2y=self.ccd2_roi2y,
                                                                                   roi2xend=self.ccd2_roi2xend,
                                                                                   roi2yend=self.ccd2_roi2yend,
                                                                                   thresh0=self.chile_window.ccd2yuzhi[0],
                                                                                   thresh1=self.chile_window.ccd2yuzhi[
                                                                                       1])  # 长度距离检测
                    print("distance is :", self.gap)
                    self.gap = round(camera2k, 15) * self.gap
                    self.cam2result = self.gap
                elif self.camera2_strategy == int(1):#制动臂
                    # 图像处理
                    self.distance, self.img2 = procCallback.cepianyizhi(numpy_image, roi1x=self.ccd2_roi1x, roi1y=self.ccd2_roi1y,roi1xend=self.ccd2_roi1xend,roi1yend=self.ccd2_roi1yend,
                                                                       roi2x = self.ccd2_roi2x, roi2y=self.ccd2_roi2y,roi2xend=self.ccd2_roi2xend,roi2yend=self.ccd2_roi2yend,
                                                                       thresh0 =self.chile_window.ccd2yuzhi[0],thresh1 = self.chile_window.ccd2yuzhi[1])  # 偏心距
                    print("虚拟pianxiju:", self.distance)
                    self.distance = round(camera2k, 15) * self.distance
                    # todo distance为轴到标定板的距离
                    # 实际偏心距离distance = distance - 实际中心距离L
                    # self.distance = self.distance - self.L
                    # 制动臂左侧到圆心距离
                    self.cam2result = self.distance
                    self.distance = round(camera2k, 15) * self.distance
                    self.cam2result = self.distance
                elif self.camera2_strategy == int(2):  # 间隙测距
                    # 图像处理
                    # todo 2022.07.22
                    self.distance, self.img2 = procCallback.measure_gear_gap(numpy_image, roi1x=self.ccd2_roi1x, roi1y=self.ccd2_roi1y,roi1xend=self.ccd2_roi1xend,roi1yend=self.ccd2_roi1yend,
                                                                       roi2x = self.ccd2_roi2x, roi2y=self.ccd2_roi2y,roi2xend=self.ccd2_roi2xend,roi2yend=self.ccd2_roi2yend,
                                                                       thresh0 =self.chile_window.ccd2yuzhi[0],thresh1 = self.chile_window.ccd2yuzhi[1])
                    print("distance2:", self.distance)
                    self.distance = round(camera2k, 15) * self.distance
                    self.cam2result = self.distance
            # todo 误差补偿
            self.cam2result = round((self.cam2result + (float(self.chile_window.lineEdit_6.text()))), 2)
            print('camera2result is ', self.cam2result)
            self.label_32.setText("相机二结果" + str(self.cam2result))
            # todo 结果检查
            if self.label_window.liheqi==int(0):
                if abs(self.cam2result) <= float(self.label_window.lineEdit_42.text()) and abs(
                        self.cam2result) >= float(self.label_window.lineEdit_31.text()):  # 检测条件构建
                    self.cam2good = self.cam2good + 1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B29'] = self.cam2good
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("OKSHOW.jpg", ))
                    self.camera2flag = True
                else:
                    # self.cam2good = self.cam2good+1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera2flag = False
            elif self.label_window.liheqi==int(1):
                if abs(self.cam2result) <= float(self.label_window.lineEdit_20.text()) and abs(self.cam2result) >= float(
                        self.label_window.lineEdit_19.text()):  # 检测条件构建
                    self.cam2good = self.cam2good + 1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B29'] = self.cam2good
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("OKSHOW.jpg", ))
                    self.camera2flag = True
                else:
                    # self.cam2good = self.cam2good+1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera2flag = False
            elif self.label_window.liheqi==int(2):
                if abs(self.cam2result) <= float(self.label_window.lineEdit_28.text()) and abs(self.cam2result) >= float(
                        self.label_window.lineEdit_21.text()):  # 检测条件构建
                    self.cam2good = self.cam2good + 1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B29'] = self.cam2good
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("OKSHOW.jpg", ))
                    self.camera2flag = True
                else:
                    # self.cam2good = self.cam2good+1
                    self.cam2goodrate = self.cam2good / self.num
                    # self.label_25.setText(str(self.cam2good))
                    # self.label_23.setText(str(self.cam2goodrate))
                    self.sheet1['B30'] = self.cam2goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_11.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera2flag = False
            # show acquired image
            # img = Image.fromarray(self.img, 'L')
            # print("imgformate is:", type(img))
            # self.cv2img2 = cv2.cvtColor(np.asarray(img), cv2.COLOR_GRAY2RGB)
            # self.cv2img2 = cv2.resize(self.cv2img2, (450, 450))
            # cv2.imshow("this is cv2 window", cv2img)
            # cv2.waitKey(0)
            self.cv2img2 = cv2.resize(self.img2, (450, 450))
            self.Qframe_daheng2 = QImage(self.cv2img2.data, self.cv2img2.shape[1], self.cv2img2.shape[0],
                                         self.cv2img2.shape[1] * 3,
                                         QImage.Format_RGB888)
            # print(Qframe)
            # pix = QPixmap(Qframe).scaled(frame.shape[1], frame.shape[0])
            # self.setPixmap(pix)
            # QRect qq(20,50,self.img.width,self.img.height)
            self.label_2.setPixmap(QPixmap.fromImage(self.Qframe_daheng2))
            print("show success")
            # print height, width, and frame ID of the acquisition image
            print("Frame ID: %d   Height: %d   Width: %d"
                  % (raw_image.get_frame_id(), raw_image.get_height(), raw_image.get_width()))
        # self.cam.stream_off()
        # self.cam.close_device()

    def camera3(self):
        global camera3k, camera3b
        self.cam3.ExposureTime.set(self.ccd3_exposure_time)
        # set gain
        self.cam3.Gain.set(10.0)
        # 流开启
        self.cam3.stream_on()
        # start data acquisition+
        time.sleep(0.2)
        self.cam3.TriggerSoftware.send_command()  # 软件触发
        time.sleep(0.5)
        num = 1
        for i in range(num):
            # get raw image
            raw_image = self.cam3.data_stream[0].get_image()
            if raw_image is None:
                print("Getting image failed.")
                continue
            # create numpy array with data from raw image
            numpy_image = raw_image.get_numpy_array()
            # print(type(numpy_image)) cv2的图片了。
            if numpy_image is None:
                continue
            # todo   检测策略的选择
            # 检测策略的选择
            # 图像处理
            if self.camera3_strategy == int(0):  # 脱水轴高度
                self.gap, self.img3 = procCallback.measure_tuoshuizhou_distance(numpy_image, roi1x=self.ccd3_roi1x,
                                                                                roi1y=self.ccd3_roi1y,
                                                                                roi1xend=self.ccd3_roi1xend,
                                                                                roi1yend=self.ccd3_roi1yend,
                                                                                roi2x=self.ccd3_roi2x,
                                                                                roi2y=self.ccd3_roi2y,
                                                                                roi2xend=self.ccd3_roi2xend,
                                                                                roi2yend=self.ccd3_roi2yend,
                                                                                thresh0=self.chile_window.ccd3yuzhi[0],
                                                                                thresh1=self.chile_window.ccd3yuzhi[
                                                                                    1])  # 长度距离检测
                print("distance is :", self.gap)
                self.gap = round(camera3k, 15) * self.gap
                self.cam3result = self.gap
            elif self.camera3_strategy == int(1):  # 制动臂
                # 图像处理
                self.distance, self.img3 = procCallback.cepianyizhi(numpy_image, roi1x=self.ccd3_roi1x,
                                                                    roi1y=self.ccd3_roi1y, roi1xend=self.ccd3_roi1xend,
                                                                    roi1yend=self.ccd3_roi1yend,
                                                                    roi2x=self.ccd3_roi2x, roi2y=self.ccd3_roi2y,
                                                                    roi2xend=self.ccd3_roi2xend,
                                                                    roi2yend=self.ccd3_roi2yend,
                                                                    thresh0=self.chile_window.ccd3yuzhi[0],
                                                                    thresh1=self.chile_window.ccd3yuzhi[1])  # 偏心距
                print("制动臂距离:", self.distance)
                self.distance = round(camera3k, 15) * self.distance
                # todo distance为轴到标定板的距离
                # 实际偏心距离distance = distance - 实际中心距离L
                # self.distance = self.distance - self.L
                # 制动臂左侧到圆心距离
                self.cam3result = self.distance
            elif self.camera3_strategy == int(2):  # 三号相机检测策略
                # 图像处理
                # todo 2022.07.22
                self.distance, self.img3 = procCallback.measure_gear_gap(numpy_image, roi1x=self.ccd3_roi1x, roi1y=self.ccd3_roi1y,roi1xend=self.ccd3_roi1xend,roi1yend=self.ccd3_roi1yend,
                                                                   roi2x = self.ccd3_roi2x, roi2y=self.ccd3_roi2y,roi2xend=self.ccd3_roi2xend,roi2yend=self.ccd3_roi2yend,
                                                                   thresh0 =self.chile_window.ccd3yuzhi[0],thresh1 = self.chile_window.ccd3yuzhi[1])
                print("distance3:", self.distance)
                self.distance = round(camera3k, 15) * self.distance
                self.cam3result = self.distance
            # todo 误差补偿
            self.cam3result = round((self.cam3result + (float(self.chile_window.lineEdit_7.text()))), 2)
            print('cam3result', self.cam3result)
            self.label_33.setText("相机三结果" + str(self.cam3result))
            # todo 结果检查
            if self.label_window.liheqi==int(0):
                if self.cam3result < float(self.label_window.lineEdit_28.text()) and self.cam3result > float(
                        self.label_window.lineEdit_21.text()):  # 检测条件构建
                    self.cam3good = self.cam3good + 1
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_21.setText(str(self.cam3good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B31'] = self.cam3good
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera3flag = True
                    self.label_12.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_12.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera3flag = False
            elif self.label_window.liheqi==int(1):
                if self.cam3result < float(self.label_window.lineEdit_20.text()) and self.cam3result > float(
                        self.label_window.lineEdit_19.text()):  # 检测条件构建
                    self.cam3good = self.cam3good + 1
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_21.setText(str(self.cam3good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B31'] = self.cam3good
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera3flag = True
                    self.label_12.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_12.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera3flag = False
            elif self.label_window.liheqi==int(2):
                if self.cam3result < float(self.label_window.lineEdit_20.text()) and self.cam3result > float(
                        self.label_window.lineEdit_19.text()):  # 检测条件构建
                    self.cam3good = self.cam3good + 1
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_21.setText(str(self.cam3good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B31'] = self.cam3good
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.camera3flag = True
                    self.label_12.setPixmap(QPixmap("OKSHOW.jpg", ))
                else:
                    self.cam3goodrate = self.cam3good / self.num
                    # self.label_24.setText(str(self.cam2good))
                    # self.label_26.setText(str(self.cam3goodrate))
                    self.sheet1['B32'] = self.cam3goodrate
                    print("parameter get")
                    md.workbook_initial.save('parameter_record.xlsx')  # 保存文档
                    print("saved done")
                    self.label_12.setPixmap(QPixmap("NGSHOW.jpg", ))
                    self.camera3flag = False
            # show acquired image
            # img = Image.fromarray(self.img, 'L')
            # print("imgformate is:", type(img))
            # self.cv2img = cv2.cvtColor(np.asarray(img), cv2.COLOR_GRAY2RGB)

            # self.cv2img = cv2.resize(self.cv2img, (450, 450))

            # cv2.imshow("this is cv2 window", cv2img)
            # cv2.waitKey(0)
            self.cv2img3 = cv2.resize(self.img3, (450, 450))
            self.Qframe_daheng3 = QImage(self.cv2img3.data, self.cv2img3.shape[1], self.cv2img3.shape[0],
                                         self.cv2img3.shape[1] * 3,
                                         QImage.Format_RGB888)
            # print(Qframe)
            # pix = QPixmap(Qframe).scaled(frame.shape[1], frame.shape[0])
            # self.setPixmap(pix)
            # QRect qq(20,50,self.img.width,self.img.height)
            self.label_3.setPixmap(QPixmap.fromImage(self.Qframe_daheng3))
            print("show success")
            # print height, width, and frame ID of the acquisition image
            print("Frame ID: %d   Height: %d   Width: %d"
                  % (raw_image.get_frame_id(), raw_image.get_height(), raw_image.get_width()))
        # self.cam.stream_off()
        # self.cam.close_device()

    def print_trig(self):
        # 获得product的打印信息
        while 1:
            time.sleep(0.1)
            self.label_30.setText(str(round(self.cam1result, 2)))
            self.label_32.setText(str(round(self.cam2result, 2)))
            self.label_33.setText(str(round(self.cam3result, 2)))
            self.serial = str(self.serial_number + 100000)
            self.serial = str(self.serial)[-4:]
            # todo only 3 camera now
            if self.camera1flag == 1 and self.camera2flag == 1 and self.camera3flag == 1:
                # todo 优化打印标签的信息
                self.serial_number += 1
                yyy = time.strftime('%y', time.localtime(time.time()))
                mmm = time.strftime('%m', time.localtime(time.time()))
                ddd = time.strftime('%d', time.localtime(time.time()))
                YYY = Year[yyy]
                MMM = Month[mmm]
                DDD = Day[ddd]
                #todo adaption print
                #todo 自动选择对应型号
                if self.label_window.liheqi == int(0):
                    # N-15.0离合器
                    self.product_infomation = self.label_window.lineEdit.text() + self.label_window.lineEdit_2.text() + self.label_window.lineEdit_3.text() + YYY + MMM + DDD + \
                                              self.serial + self.label_window.lineEdit_7.text() + self.label_window.lineEdit_6.text() + self.label_window.lineEdit_5.text() + self.label_window.lineEdit_11.text() + self.label_window.lineEdit_8.text() + self.label_window.lineEdit_10.text() + self.starttime + "/" + str(
                        time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))) + "/" + \
                                              self.label_window.lineEdit_23.text() + '-' + str(
                        (format(self.cam3result,
                                '.2f'))) + '-' + self.label_window.lineEdit_28.text() + '-' + self.label_window.lineEdit_21.text() + '/' + self.label_window.lineEdit_22.text() + '-' + str(
                        format(self.cam1result,
                               '.2f')) + '-' + self.label_window.lineEdit_20.text() + '-' + self.label_window.lineEdit_19.text() + '/' + \
                                              self.label_window.lineEdit_44.text() + '-' + str(
                        format(self.cam2result,
                               '.2f')) + '-' + self.label_window.lineEdit_42.text() + '-' + self.label_window.lineEdit_31.text() + "/"
                    print(self.product_infomation)

                elif self.label_window.liheqi == int(1) :
                    # N-5.25离合器
                    self.product_infomation = self.label_window.lineEdit.text() + self.label_window.lineEdit_2.text() + self.label_window.lineEdit_3.text() + YYY + MMM + DDD + \
                                              self.serial + self.label_window.lineEdit_7.text() + self.label_window.lineEdit_6.text() + self.label_window.lineEdit_5.text() + self.label_window.lineEdit_11.text() + self.label_window.lineEdit_8.text() + self.label_window.lineEdit_10.text() + self.starttime + "/" + str(
                        time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))) + "/" + \
                                              self.label_window.lineEdit_23.text() + '-' + str(
                        (format(self.cam3result,
                                '.2f'))) + '-' + self.label_window.lineEdit_28.text() + '-' + self.label_window.lineEdit_21.text() + '/' + self.label_window.lineEdit_22.text() + '-' + str(
                        format(self.cam2result,
                               '.2f')) + '-' + self.label_window.lineEdit_20.text() + '-' + self.label_window.lineEdit_19.text() + '/' + \
                                              self.label_window.lineEdit_44.text() + '-' + str(
                        format(self.cam1result,
                               '.2f')) + '-' + self.label_window.lineEdit_42.text() + '-' + self.label_window.lineEdit_31.text() + "/"
                    print(self.product_infomation)
                elif self.label_window.liheqi == int(2):
                    # N-11.0离合器
                    self.product_infomation = self.label_window.lineEdit.text() + self.label_window.lineEdit_2.text() + self.label_window.lineEdit_3.text() + YYY + MMM + DDD + \
                                              self.serial + self.label_window.lineEdit_7.text() + self.label_window.lineEdit_6.text() + self.label_window.lineEdit_5.text() + self.label_window.lineEdit_11.text() + self.label_window.lineEdit_8.text() + self.label_window.lineEdit_10.text() + self.starttime + "/" + str(
                        time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))) + "/" + \
                                              self.label_window.lineEdit_23.text() + '-' + str(
                        (format(self.cam2result,
                                '.2f'))) + '-' + self.label_window.lineEdit_28.text() + '-' + self.label_window.lineEdit_21.text() + '/' + self.label_window.lineEdit_22.text() + '-' + str(
                        format(self.cam3result,
                               '.2f')) + '-' + self.label_window.lineEdit_20.text() + '-' + self.label_window.lineEdit_19.text() + '/' + \
                                              self.label_window.lineEdit_44.text() + '-' + str(
                        format(self.cam1result,
                               '.2f')) + '-' + self.label_window.lineEdit_42.text() + '-' + self.label_window.lineEdit_31.text() +"/"
                    print(self.product_infomation)
                self.label_34.setText(self.product_infomation)
                self.print_label(self.product_infomation)
                self.camera1flag = False
                self.camera2flag = False
                self.camera3flag = False
                # recor data
                # todo
                data_record(self.product_infomation)
                # record parameter data into parameter_record.xlsx
                self.sheet1['B1'] = self.serial_number  # 在该行更新序列号
                self.workbook_initial.save('parameter_record.xlsx')  # 保存文档
            else:
                pass

    def print_label(self, Qinfo="need more information"):  # need string
        data = Qinfo
        img_file = 'code.jpg'
        qr = qrcode.QRCode(version=5,
                           # error_correction=qrcode.constants.ERROR_CORRECT_H,
                           box_size=12, border=8)
        qr.add_data(data)
        # qr.make(fit=True)
        img = qr.make_image()
        img.save(img_file)
        img = cv2.imread(img_file)
        img = img[80:585, 80:585]
        img = cv2.resize(img, (220, 220))

        #############构建图片与文字#######################
        newImg = np.zeros((350, 480, 3), dtype=np.uint8)
        newImg[:] = [255, 255, 255]
        fonts = self.fonts
        # font=cv2.FONT_HERSHEY_PLAIN
        # todo 不同型号qr二维码排列方式不一样
        fontsize = 1
        # todo 字体可以选择，宋体，黑体等其他字体
        if self.label_window.liheqi == int(0):
            # N-15.0离合器
            newImg[20:240, 4:224] = img
            cv2.imwrite("img_code.jpg", newImg)
            im = Image.open('img_code.jpg')
            draw = ImageDraw.Draw(im)
            fnt = ImageFont.truetype(fonts, 35)  # 字体与大小
            draw.text((235, 20),'SX(' + str(self.label_window.lineEdit_3.text()) + ')', fill='black', font=fnt)
            draw.text((235, 80), str(time.strftime('%Y%m%d', time.localtime(time.time()))), fill='black', font=fnt)
            draw.text((235, 140), str(self.zubie) + self.serial, fill='black', font=fnt)
            draw.text((235, 200), str(self.label_window.lineEdit_2.text()), fill='black', font=fnt)
            im.save("savedcode.jpg")

        elif self.label_window.liheqi == int(1):
            # N-5.25离合器
            newImg[20:240, 4:224] = img
            cv2.imwrite("img_code.jpg", newImg)
            im = Image.open('img_code.jpg')
            draw = ImageDraw.Draw(im)
            fnt = ImageFont.truetype(fonts, 35)  # 字体与大小
            draw.text((235, 20),'SX(' + str(self.label_window.lineEdit_3.text()) + ')', fill='black', font=fnt)
            draw.text((235, 80), str(time.strftime('%Y%m%d', time.localtime(time.time()))), fill='black', font=fnt)
            draw.text((235, 140), str(self.zubie) + self.serial, fill='black', font=fnt)
            draw.text((235, 200), str(self.label_window.lineEdit_2.text()), fill='black', font=fnt)
            im.save("savedcode.jpg")
        elif self.label_window.liheqi == int(2):
            # NBP-11.0离合器
            newImg[20:240, 4:224] = img
            cv2.imwrite("img_code.jpg", newImg)
            im = Image.open('img_code.jpg')
            draw = ImageDraw.Draw(im)
            fnt = ImageFont.truetype(fonts, 35)  # 字体与大小
            # fnt2 = ImageFont.truetype('C:/Windows/Fonts/STFANGSO.TTF', 40)  # 仿宋
            # fnt3 = ImageFont.truetype('C:/Windows/Fonts/STXIHEI.TTF', 50)  # 细黑
            # fnt4 = ImageFont.truetype('C:/Windows/Fonts/STFANGSO.TTF', 50)  # 仿宋
            draw.text((235, 20), 'SX(' + str(self.label_window.lineEdit_3.text()) + ')', fill='black', font=fnt)
            draw.text((235, 80), str(self.label_window.lineEdit_2.text()), fill='black', font=fnt)
            draw.text((235, 140), str(time.strftime('%Y%m%d', time.localtime(time.time()))), fill='black', font=fnt)
            draw.text((235, 200), str(self.zubie) + self.serial, fill='black', font=fnt)
            # im.show()
            im.save("savedcode.jpg")
            '''
            cv2.putText(newImg, 'SX(' + str(self.label_window.lineEdit_3.text()) + ')', (236, 42), font, fontsize,
                        (0, 0, 0), 3)
            cv2.putText(newImg, str(self.label_window.lineEdit_2.text()), (236, 97), font, fontsize, (0, 0, 0), 3)
            cv2.putText(newImg, str(time.strftime('%Y%m%d', time.localtime(time.time()))), (236, 151), font, fontsize,
                        (0, 0, 0),3)
            cv2.putText(newImg, str(self.zubie) + self.serial, (236, 208), font, fontsize, (0, 0, 0),3)

            newImg = cv2.imwrite("savedcode.jpg", newImg)
            '''
        else:
            pass

        #################打印机连接与打印##############################

        # 物理宽度、高度
        PHYSICALWIDTH = 90
        PHYSICALHEIGHT = 60
        # 物理偏移位置

        PHYSICALOFFSETX = 0
        PHYSICALOFFSETY = 3
        printer_name = win32print.GetDefaultPrinter()
        # 选择图片路径
        file_name = "savedcode.jpg"

        hDC = win32ui.CreateDC()
        hDC.CreatePrinterDC(printer_name)
        printer_size = hDC.GetDeviceCaps(PHYSICALWIDTH), hDC.GetDeviceCaps(PHYSICALHEIGHT)
        printer_margins = hDC.GetDeviceCaps(PHYSICALOFFSETX), hDC.GetDeviceCaps(PHYSICALOFFSETY)
        # 打开图片
        bmp = Image.open(file_name)

        print(bmp.size)
        ratios = [1.0 * 1754 / bmp.size[0], 1.0 * 1240 / bmp.size[1]]
        scale = min(ratios)
        print(ratios)
        print(scale)
        hDC.StartDoc(file_name)
        hDC.StartPage()

        dib = ImageWin.Dib(bmp)

        scaled_width, scaled_height = [int(scale * i) for i in bmp.size]
        print(scaled_width, scaled_height)
        x1 = int((printer_size[0] - scaled_width) / 2)
        y1 = int((printer_size[1] - scaled_height) / 2)
        # 横向位置坐标
        x1 = 0
        # 竖向位置坐标
        y1 = 0
        # 4倍为自适应图片实际尺寸打印
        x2 = x1 + bmp.size[0] * 0.8
        y2 = y1 + bmp.size[1] * 0.8
        dib.draw(hDC.GetHandleOutput(), (x1, y1, int(x2), int(y2)))

        hDC.EndPage()
        hDC.EndDoc()
        hDC.DeleteDC()
        print("finish print")

    def selectionchange(self, i):
        print("Items in the list are :")
        for count in range(self.comboBox.count()):
            print(self.comboBox.itemText(count))  # Displays text belonging to specific index
        print("Current index", i, "selection changed ", self.comboBox.currentText())
        if i == 0:
            print("choose 脱水轴（大油封）高度")
            self.camera1_strategy = int(0)
        elif i == 1:
            print("choose 制动臂间距")
            self.camera1_strategy = int(1)
        elif i == 2:
            print("choose 连接轴与传矩轴套间隙")
            self.camera1_strategy = int(2)

    def selectionchange_2(self, i):
        print("Items in the list are :")
        for count in range(self.comboBox_2.count()):
            print(self.comboBox_2.itemText(count))  # Displays text belonging to specific index
        print("Current index", i, "selection changed ", self.comboBox_2.currentText())
        if i == 0:
            print("choose 脱水轴（大油封）高度")
            self.camera2_strategy = int(0)
        elif i == 1:
            print("choose 制动臂间距")
            self.camera2_strategy = int(1)
        elif i == 2:
            print("choose 连接轴与传矩轴套间隙")
            self.camera2_strategy = int(2)

    def selectionchange_3(self, i):
        print("Items in the list are :")
        for count in range(self.comboBox_3.count()):
            print(self.comboBox_3.itemText(count))  # Displays text belonging to specific index
        print("Current index", i, "selection changed ", self.comboBox_3.currentText())
        if i == 0:
            print("choose 脱水轴（大油封）高度")
            self.camera3_strategy = int(0)
        elif i == 1:
            print("choose 制动臂间距")
            self.camera3_strategy = int(1)
        elif i == 2:
            print("choose 连接轴与传矩轴套间隙")
            self.camera3_strategy = int(2)

    def paintEvent(self, a0: QtGui.QPaintEvent):
        # 一直后台刷新
        self.label_6.setText(str(time.strftime('%Y-%m-%d %H:%M', time.localtime(time.time()))))


###########################主程序Main Proc############################
if __name__ == '__main__':
    app = QApplication(sys.argv)
    md = MainCode()
    md.show()
    sys.exit(app.exec_())
#################
